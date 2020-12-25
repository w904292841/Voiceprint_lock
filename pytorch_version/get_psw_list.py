import os
from openpyxl import load_workbook
import numpy as np

fa_path = "F:/文件/Voiceprint_recognition_rewrite-master/a_real/"
list_file_name = "password.xlsx"

workbook = load_workbook(fa_path + list_file_name)
booksheet = workbook.active

#获取sheet页的行数据
rows = booksheet.rows
#获取sheet页的列数据
columns = booksheet.columns

i = 0
# 迭代所有的行

file_name = []
password = []

for row in rows:
    i = i + 1
    line = [col.value for col in row]
    file_name.append(booksheet.cell(row=i, column=1).value)              #获取第i行1 列的数据
    password.append(booksheet.cell(row=i, column=2).value)               #获取第i行 2 列的数据

file_name = np.array(file_name)
password = np.array(password)

np.save(fa_path + "file_name", file_name)
np.save(fa_path + "password", password)